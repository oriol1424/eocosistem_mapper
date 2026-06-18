import io
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

CATEGORIAS = [
    "Empresas privadas",
    "Academia",
    "Administración pública",
    "Sociedad civil organizada"
]

COLORES_CABECERA = {
    "Empresas privadas":        "1D6FA8",
    "Academia":                 "6B4FA8",
    "Administración pública":   "1A7A4A",
    "Sociedad civil organizada":"A85A1A"
}

COLUMNAS_POR_CATEGORIA = {
    "Empresas privadas":         ["Nombre","Tipo","Sector","Descripción","Web","Dirección","Teléfono","Ubicación"],
    "Academia":                  ["Nombre","Tipo","Sector","Descripción","Web","Dirección","Teléfono","Ubicación"],
    "Administración pública":    ["Nombre","Tipo","Descripción","Web","Dirección","Teléfono","Horarios","Ubicación"],
    "Sociedad civil organizada": ["Nombre","Tipo","Sector","Descripción","Web","Dirección","Teléfono","Ubicación"],
}

CAMPOS_POR_CATEGORIA = {
    "Empresas privadas":         ["nombre","tipo","sector","descripcion","web","direccion","contacto","ubicacion"],
    "Academia":                  ["nombre","tipo","sector","descripcion","web","direccion","contacto","ubicacion"],
    "Administración pública":    ["nombre","tipo","descripcion","web","direccion","contacto","horarios","ubicacion"],
    "Sociedad civil organizada": ["nombre","tipo","sector","descripcion","web","direccion","contacto","ubicacion"],
}


def _borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _escribir_hoja_datos(ws, categoria, actores, zona):
    color = COLORES_CABECERA[categoria]
    columnas = COLUMNAS_POR_CATEGORIA[categoria]
    campos = CAMPOS_POR_CATEGORIA[categoria]
    borde = _borde_fino()

    ws.merge_cells(f"A1:{get_column_letter(len(columnas))}1")
    c = ws["A1"]
    c.value = f"{categoria} — {zona}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[2].height = 22

    for row_idx, actor in enumerate(actores, 3):
        bg = "F7F7F7" if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, campo in enumerate(campos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=actor.get(campo, "") or "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = borde

    anchos = [30, 22, 22, 50, 35, 35, 20, 20]
    for col_idx, ancho in enumerate(anchos[:len(columnas)], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.freeze_panes = "A3"
    if actores:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columnas))}{len(actores)+2}"
    if not actores:
        ws.cell(row=3, column=1, value="No se encontraron actores para esta categoría.")


def _celda_titulo(ws, row, texto, color):
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32


def _celda_subtitulo(ws, row, col, texto, color):
    c = ws.cell(row=row, column=col, value=texto)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _metrica(ws, row, col, label, valor, color):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(size=10, color="666666")
    vc = ws.cell(row=row+1, column=col, value=valor)
    vc.font = Font(bold=True, size=18, color=color)
    vc.alignment = Alignment(horizontal="center")


def _escribir_tabla_conteo(ws, start_row, start_col, titulo, conteo, color, max_items=8):
    ws.cell(row=start_row, column=start_col, value=titulo).font = Font(bold=True, size=10, color="FF" + color)
    items = conteo.most_common(max_items)
    for i, (label, count) in enumerate(items):
        ws.cell(row=start_row+1+i, column=start_col, value=label or "Sin clasificar")
        ws.cell(row=start_row+1+i, column=start_col+1, value=count)
    return start_row + 1 + len(items)


def _grafico_barras(ws, data_row_start, data_row_end, col_label, col_val, titulo, anchor, color_hex):
    chart = BarChart()
    chart.type = "bar"
    chart.title = titulo
    chart.style = 10
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.width = 14
    chart.height = 10
    chart.grouping = "clustered"

    data = Reference(ws, min_col=col_val, min_row=data_row_start, max_row=data_row_end)
    cats = Reference(ws, min_col=col_label, min_row=data_row_start, max_row=data_row_end)
    chart.add_data(data)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = color_hex
    chart.series[0].title = None
    chart.legend = None
    ws.add_chart(chart, anchor)


def _grafico_pie(ws, data_row_start, data_row_end, col_label, col_val, titulo, anchor):
    chart = PieChart()
    chart.title = titulo
    chart.style = 10
    chart.width = 14
    chart.height = 10

    data = Reference(ws, min_col=col_val, min_row=data_row_start, max_row=data_row_end)
    cats = Reference(ws, min_col=col_label, min_row=data_row_start, max_row=data_row_end)
    chart.add_data(data)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def _dashboard_empresas(ws, actores, zona, color):
    _celda_titulo(ws, 1, f"Dashboard Empresas Privadas — {zona}", color)

    total = len(actores)
    _metrica(ws, 3, 1, "Total empresas", total, "FF" + color)

    sectores = Counter(a.get("sector", "") or "Sin sector" for a in actores)
    tipos = Counter(a.get("tipo", "") or "Sin tipo" for a in actores)

    # Tabla sectores (col A)
    _celda_subtitulo(ws, 6, 1, "Por sector", color)
    row_end_s = _escribir_tabla_conteo(ws, 6, 1, "Por sector", sectores, color)

    # Tabla tipos (col D)
    _celda_subtitulo(ws, 6, 4, "Por tipo", color)
    row_end_t = _escribir_tabla_conteo(ws, 6, 4, "Por tipo", tipos, color)

    max_row = max(row_end_s, row_end_t)

    # Gráfico barras sectores
    if len(sectores) > 0:
        _grafico_barras(ws, 7, 7+min(len(sectores)-1,7), 1, 2, "Sectores más relevantes", f"A{max_row+2}", color)

    # Gráfico pie tipos
    if len(tipos) > 0:
        _grafico_pie(ws, 7, 7+min(len(tipos)-1,7), 4, 5, "Distribución por tipo", f"J{max_row+2}")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10


def _dashboard_academia(ws, actores, zona, color):
    _celda_titulo(ws, 1, f"Dashboard Academia — {zona}", color)

    total = len(actores)
    _metrica(ws, 3, 1, "Total centros", total, "FF" + color)

    tipos = Counter(a.get("tipo", "") or "Sin tipo" for a in actores)

    universidades = sum(1 for a in actores if "univers" in (a.get("tipo","") or "").lower())
    fp = sum(1 for a in actores if any(x in (a.get("tipo","") or "").lower() for x in ["fp","formación prof","vocacional","técni"]))
    master = sum(1 for a in actores if any(x in (a.get("tipo","") or "").lower() for x in ["máster","master","posgrado","postgrado"]))
    doctorado = sum(1 for a in actores if any(x in (a.get("tipo","") or "").lower() for x in ["doctor","phd","tesis"]))
    investigacion = sum(1 for a in actores if any(x in (a.get("tipo","") or "").lower() for x in ["investig","research","i+d"]))
    otros = total - universidades - fp - master - doctorado - investigacion

    nivel_data = [
        ("Universidades", universidades),
        ("Formación Profesional", fp),
        ("Másteres / Posgrado", master),
        ("Doctorado / PhD", doctorado),
        ("Investigación / I+D", investigacion),
        ("Otros centros", max(otros, 0)),
    ]

    _celda_subtitulo(ws, 6, 1, "Por nivel formativo", color)
    for i, (label, val) in enumerate(nivel_data):
        ws.cell(row=7+i, column=1, value=label)
        ws.cell(row=7+i, column=2, value=val)

    _celda_subtitulo(ws, 6, 4, "Por tipo detallado", color)
    row_end = _escribir_tabla_conteo(ws, 6, 4, "Por tipo", tipos, color)

    if total > 0:
        _grafico_pie(ws, 7, 7+len(nivel_data)-1, 1, 2, "Distribución por nivel formativo", "A16")
        _grafico_barras(ws, 7, 7+min(len(tipos)-1,7), 4, 5, "Tipos de centro", "J16", color)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10


def _dashboard_administracion(ws, actores, zona, color):
    _celda_titulo(ws, 1, f"Dashboard Administración Pública — {zona}", color)

    total = len(actores)
    _metrica(ws, 3, 1, "Total entidades", total, "FF" + color)

    tipos = Counter(a.get("tipo", "") or "Sin tipo" for a in actores)

    aytos = sum(1 for a in actores if "ayunt" in (a.get("tipo","") or "").lower() or "ayunt" in (a.get("nombre","") or "").lower())
    emp_pub = sum(1 for a in actores if "empresa púb" in (a.get("tipo","") or "").lower())
    agencias = sum(1 for a in actores if "agencia" in (a.get("tipo","") or "").lower())
    servicios = sum(1 for a in actores if "servicio" in (a.get("tipo","") or "").lower())
    otros = total - aytos - emp_pub - agencias - servicios

    tipo_data = [
        ("Ayuntamientos", aytos),
        ("Empresas públicas", emp_pub),
        ("Agencias", agencias),
        ("Servicios sociales", servicios),
        ("Otros organismos", max(otros, 0)),
    ]

    _celda_subtitulo(ws, 6, 1, "Por tipo de entidad", color)
    for i, (label, val) in enumerate(tipo_data):
        ws.cell(row=7+i, column=1, value=label)
        ws.cell(row=7+i, column=2, value=val)

    _celda_subtitulo(ws, 6, 4, "Detalle por tipo", color)
    _escribir_tabla_conteo(ws, 6, 4, "Detalle", tipos, color)

    if total > 0:
        _grafico_barras(ws, 7, 7+len(tipo_data)-1, 1, 2, "Distribución de entidades públicas", "A14", color)
        _grafico_pie(ws, 7, 7+min(len(tipos)-1,7), 4, 5, "Detalle por tipo", "J14")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10


def _dashboard_sociedad(ws, actores, zona, color):
    _celda_titulo(ws, 1, f"Dashboard Sociedad Civil Organizada — {zona}", color)

    total = len(actores)
    _metrica(ws, 3, 1, "Total organizaciones", total, "FF" + color)

    tipos = Counter(a.get("tipo", "") or "Sin tipo" for a in actores)
    sectores = Counter(a.get("sector", "") or "Sin sector" for a in actores)

    ongs = sum(1 for a in actores if "ong" in (a.get("tipo","") or "").lower())
    fundaciones = sum(1 for a in actores if "fundac" in (a.get("tipo","") or "").lower())
    asociaciones = sum(1 for a in actores if "asoc" in (a.get("tipo","") or "").lower())
    cooperativas = sum(1 for a in actores if "cooper" in (a.get("tipo","") or "").lower())
    vecinales = sum(1 for a in actores if "vecin" in (a.get("tipo","") or "").lower())
    otros = total - ongs - fundaciones - asociaciones - cooperativas - vecinales

    forma_data = [
        ("ONGs", ongs),
        ("Fundaciones", fundaciones),
        ("Asociaciones", asociaciones),
        ("Cooperativas", cooperativas),
        ("Vecinales", vecinales),
        ("Otras", max(otros, 0)),
    ]

    _celda_subtitulo(ws, 6, 1, "Por forma jurídica", color)
    for i, (label, val) in enumerate(forma_data):
        ws.cell(row=7+i, column=1, value=label)
        ws.cell(row=7+i, column=2, value=val)

    _celda_subtitulo(ws, 6, 4, "Por ámbito temático", color)
    _escribir_tabla_conteo(ws, 6, 4, "Ámbito", sectores, color)

    if total > 0:
        _grafico_pie(ws, 7, 7+len(forma_data)-1, 1, 2, "Distribución por forma jurídica", "A16")
        _grafico_barras(ws, 7, 7+min(len(sectores)-1,7), 4, 5, "Ámbitos temáticos", "J16", color)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 10


DASHBOARD_FUNCIONES = {
    "Empresas privadas":         _dashboard_empresas,
    "Academia":                  _dashboard_academia,
    "Administración pública":    _dashboard_administracion,
    "Sociedad civil organizada": _dashboard_sociedad,
}


def exportar_excel(resultados: dict, zona: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for categoria in CATEGORIAS:
        actores = resultados.get(categoria, [])
        color = COLORES_CABECERA[categoria]

        # Hoja de datos
        ws_data = wb.create_sheet(title=categoria[:31])
        _escribir_hoja_datos(ws_data, categoria, actores, zona)

        # Hoja de dashboard
        dash_title = f"📊 {categoria[:22]}"
        ws_dash = wb.create_sheet(title=dash_title)
        DASHBOARD_FUNCIONES[categoria](ws_dash, actores, zona, color)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
